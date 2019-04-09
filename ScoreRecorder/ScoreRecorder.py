# Hold the person's data
from openpyxl import load_workbook
import shelve

class PersonRecord:
    def __init__(self, name="None", idkey="0", classes="None", scores={}):
        '''
        Score={'w1':0,'t1':0,'m'=0,'f'=0} where w1:week 1, t1: test 1, 
        m: middle term exam, f: final exam
        '''
        self.name = name
        self.idkey = idkey
        self.classes = classes
        self.scores = scores

    def setScore(self, key, value):
        '''
        If there is already a data, it will be replaced without waring, otherwise creative 
        a new data
        '''
        self.scores[key] = value
    
    def getStrFormat(self):
        str="%s\t%s\t%s\n" % ("name","idkey","class")
        n=1
        for key in self.scores.keys():
            str+="%s\t" % key
            if n % 6 ==0:
                str+="\n"
            n+=1
        if str[-1]=='\n':
            str=str[0:-1]
        return str

    def __str__(self):
        str="%s\t%s\t%s\n" % (self.name,self.idkey,self.classes)
        n=1
        for key in self.scores.keys():
            str+="%5s: %s\t" % (key,self.scores[key])
            if n % 6 ==0:
                str+="\n"
            n+=1
        if str[-1]=='\n':
            str=str[0:-1]
        return str

class Data:
    def __init__(self,dataName="Data"):
        self.dataFile=dataName
        self.db=shelve.open(self.dataFile)
        if 'ExcelInit' in self.db.keys():
            obj=self.db['ExcelInit']
            self.excelFileName=obj[0]
            self.rowBegin=obj[1]
            self.rowEnd=obj[2]
            self.columnIndex=list(obj[3:])
            self.validScoresIndex=["w%d" % (x-self.columnIndex[3][0]+1) for x in self.columnIndex[3]]
            self.validScoresIndex.extend(["t%d" % (x-self.columnIndex[4][0]+1) for x in self.columnIndex[4]])
            self.validScoresIndex.extend(["m","f"])

    
    def addPerson(self,person):
        self.db[person.idkey]=person
    
    
    def readXlsxFromString(self,s,mutex):
        import ast
        if len(s) < 10:
            return False
        filename=s[0]
        rowBegin=int(s[1])
        rowEnd=int(s[2])
        name=int(s[3])
        idkey=int(s[4])
        classes=int(s[5])
        # safely evaluate to list
        weekScores=ast.literal_eval(s[6])
        testScores=ast.literal_eval(s[7])
        mScores=int(s[8])
        fScores=int(s[9])
        self.readXlsx(filename,rowBegin,rowEnd,name,idkey,classes,weekScores,testScores,mScores,fScores)
        mutex.acquire()
        return True

    
    def readXlsx(self,filename,rowBegin,rowEnd,name,idkey,classes,weekScores,testScores,mScores,fScores):
        '''
        weekScores=[1,2,3]
        d.readXlsx("02.xlsx",6,167,4,3,5,[x for x in range(11,29)],[x for x in range(29,33)],8,9)
        Openpyxl's column index starts at 1.
        '''

        self.db['ExcelInit']=(filename,rowBegin,rowEnd,name,idkey,classes,weekScores,testScores,mScores,fScores)
        
        obj=self.db['ExcelInit']
        self.excelFileName=obj[0]
        self.rowBegin=obj[1]
        self.rowEnd=obj[2]
        self.columnIndex=list(obj[3:])

        wb=load_workbook(filename,True)
        ws=wb[wb.sheetnames[0]]
        for i in range(rowBegin,rowEnd):
            pname=ws.cell(i,name).value
            # Delete blank
            pname=pname.split(' ',1)[0]
            pidkey=ws.cell(i,idkey).value
            pidkey=pidkey.split(' ',1)[0]
            pclasses=ws.cell(i,classes).value
            pclasses=pclasses.split(' ',1)[0]
            person=PersonRecord(pname,pidkey,pclasses)
            n=0
            for j in weekScores:
                n+=1
                person.setScore("w%d" % n,ws.cell(i,j).value)
            n=0
            for j in testScores:
                n+=1
                person.setScore("t%d" % n,ws.cell(i,j).value)
            person.setScore('m',ws.cell(i,mScores).value)
            person.setScore('f',ws.cell(i,fScores).value)
            #print(person)
            self.addPerson(person)

        self.save()
    
    def writeXlsx(self,filename="backup.xlsx"):
        '''
        create a new file <filename>
        '''
        wb=load_workbook(self.excelFileName,False)
        ws=wb[wb.sheetnames[0]]
        for i in range(self.rowBegin,self.rowEnd):
            idd=ws.cell(i,self.columnIndex[1]).value
            idd=idd.split(' ')[0]
            person=self.getPerson(idd)
            ws.cell(i,self.columnIndex[0],person.name)
            ws.cell(i,self.columnIndex[1],person.idkey)
            ws.cell(i,self.columnIndex[2],person.classes)
            n=0
            for j in self.columnIndex[3]:
                n+=1
                ws.cell(i,j,person.scores["w%d" % n])
            n=0
            for j in self.columnIndex[4]:
                n+=1
                ws.cell(i,j,person.scores["t%d" % n])
            ws.cell(i,self.columnIndex[5],person.scores['m'])
            ws.cell(i,self.columnIndex[6],person.scores['f'])
        
        wb.save(filename)
    
    def displayAll(self):
        strs=""
        for index in self.db.keys():
            if index == 'ExcelInit':
                continue
            strs+=self.db[index].__str__()+'\n\n'
        return strs
    
    def getPerson(self,idKey):
        p=self.db[idKey]
        return p

    def updatePersonScore(self,idKey,scoreKey,score):
        '''
        scoreKey={'w1':0,'t1':0,'m'=0,'f'=0}
        '''
        mobject=self.db[idKey]

        if scoreKey not in mobject.scores.keys():
            raise KeyError('No such scoreKey!')
        
        mobject.scores[scoreKey]=score
        self.db[idKey]=mobject
        
    
    def save(self):
        self.db.close()
        self.db=shelve.open(self.dataFile)
    
    def __del__(self):
        self.db.close()
    
    # wipe all things 
    # see more in https://stackoverflow.com/questions/17341411/how-do-i-make-shelve-file-empty-in-python
    def wipeAllData(self):
        self.db.close()
        self.db=shelve.open(self.dataFile,flag='n')


if __name__=="__main__":
    d=Data()
    #d.readXlsx("02.xlsx",6,167,4,3,5,[x for x in range(11,29)],[x for x in range(29,33)],8,9)
    # d.updatePersonScore('2014141211074','m',100)
    # d.updatePersonScore('2014141211074','w1',100)
    # d.updatePersonScore('2014141211074','t1',100)
    # print(d.getPerson('2014141211074'))
    # d.writeXlsx("backup.xlsx")
    # d.wipeAllData()
    # a=[key for key in d.db.keys() if "003" in key]
    # print(a)
    # for key in d.db.keys():
    #     print(d.getPerson(key))
    p=d.getPerson('2015141223079')
    # name =张嘉珊
    print(p.getStrFormat())
    print(p)