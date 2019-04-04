# Hold the person's data
from openpyxl import load_workbook
import shelve

class PersonRecord:
    def __init__(self, name="None", idkey="0", classes="None", scores={}):
        '''
        Score={'w1':0,'t1':0,'m'=0,'f'=0} where w1:week 1, t1: test 1, 
        m: middle term exam, f: final exam
        '''
        self.name = name
        self.idkey = idkey
        self.classes = classes
        self.scores = scores

    def setScore(self, key, value):
        '''
        If there is already a data, it will be replaced without waring, otherwise creative 
        a new data
        '''
        self.scores[key] = value

    def __str__(self):
        return "%s\t%s\t\n%s" % (self.idkey, self.name, self.scores)

class Data:
    def __init__(self):
        self.dataFile='Data'
        self.db=shelve.open(self.dataFile)
        if 'ExcelInit' in self.db.keys():
            obj=self.db['ExcelInit']
            self.excelFileName=obj[0]
            self.rowBegin=obj[1]
            self.rowEnd=obj[2]
            self.columnIndex=list(obj[3:])
    
    def addPerson(self,person):
        self.db[person.idkey]=person
    
    def readXlsx(self,filename,rowBegin,rowEnd,name,idkey,classes,weekScores,testScores,mScores,fScores):
        '''
        weekScores=[1,2,3]
        d.readXlsx("02.xlsx",6,167,4,3,5,[x for x in range(11,29)],[x for x in range(29,33)],8,9)
        Openpyxl's column index starts at 1.
        '''

        self.db['ExcelInit']=(filename,rowBegin,rowEnd,name,idkey,classes,weekScores,testScores,mScores,fScores)
        
        obj=self.db['ExcelInit']
        self.excelFileName=obj[0]
        self.rowBegin=obj[1]
        self.rowEnd=obj[2]
        self.columnIndex=list(obj[3:])

        wb=load_workbook(filename,True)
        ws=wb[wb.sheetnames[0]]
        for i in range(rowBegin,rowEnd):
            pname=ws.cell(i,name).value
            # Delete blank
            pname=pname.split(' ',1)[0]
            pidkey=ws.cell(i,idkey).value
            pidkey=pidkey.split(' ',1)[0]
            pclasses=ws.cell(i,classes).value
            pclasses=pclasses.split(' ',1)[0]
            person=PersonRecord(pname,pidkey,pclasses)
            n=0
            for j in weekScores:
                n+=1
                person.setScore("w%d" % n,ws.cell(i,j).value)
            n=0
            for j in testScores:
                n+=1
                person.setScore("t%d" % n,ws.cell(i,j).value)
            person.setScore('m',ws.cell(i,mScores).value)
            person.setScore('f',ws.cell(i,fScores).value)
            #print(person)
            self.addPerson(person)

        self.save()
    
    def writeXlsx(self,filename):
        '''
        create a new file <filename>
        '''
        wb=load_workbook(self.excelFileName,False)
        ws=wb[wb.sheetnames[0]]
        for i in range(self.rowBegin,self.rowEnd):
            idd=ws.cell(i,self.columnIndex[1]).value
            idd=idd.split(' ')[0]
            person=self.getPerson(idd)
            ws.cell(i,self.columnIndex[0],person.name)
            ws.cell(i,self.columnIndex[1],person.idkey)
            ws.cell(i,self.columnIndex[2],person.classes)
            n=0
            for j in self.columnIndex[3]:
                n+=1
                ws.cell(i,j,person.scores["w%d" % n])
            n=0
            for j in self.columnIndex[4]:
                n+=1
                ws.cell(i,j,person.scores["t%d" % n])
            ws.cell(i,self.columnIndex[5],person.scores['m'])
            ws.cell(i,self.columnIndex[6],person.scores['f'])
        
        wb.save(filename)
    
    def displayAll(self):
        for index in self.db.keys():
            print(self.db[index])
        return 
    
    def getPerson(self,idKey):
        p=self.db[idKey]
        return p

    def updatePersonScore(self,idKey,scoreKey,score):
        '''
        scoreKey={'w1':0,'t1':0,'m'=0,'f'=0}
        '''
        mobject=self.db[idKey]

        if scoreKey not in mobject.scores.keys():
            raise KeyError('No such scoreKey!')
        
        mobject.scores[scoreKey]=score
        self.db[idKey]=mobject
        
    
    def save(self):
        self.db.close()
        self.db=shelve.open(self.dataFile)
    
    def __del__(self):
        self.db.close()

d=Data()
#d.readXlsx("02.xlsx",6,167,4,3,5,[x for x in range(11,29)],[x for x in range(29,33)],8,9)
d.updatePersonScore('2014141211074','m',100)
d.updatePersonScore('2014141211074','w1',100)
d.updatePersonScore('2014141211074','t1',100)
print(d.getPerson('2014141211074'))
d.writeXlsx("backup.xlsx")

